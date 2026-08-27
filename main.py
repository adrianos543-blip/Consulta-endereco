import os
import sys
import traceback
import unicodedata
from collections import defaultdict

from kivy.app import App
from kivy.clock import Clock
from kivy.lang import Builder
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.factory import Factory
from kivy.metrics import dp

# ==================================================
# LOG DE ERROS - grava num arquivo que dá pra abrir
# ==================================================
PASTA_LOG = '/storage/emulated/0/Download'
ARQUIVO_LOG = os.path.join(PASTA_LOG, 'estoque_app_erro.txt')


def registrar_erro(origem, excecao):
    """Grava o erro num arquivo de texto no Download, com data/hora."""
    try:
        import datetime
        agora = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        texto = (
            f'\n===== ERRO em {origem} - {agora} =====\n'
            + ''.join(traceback.format_exception(type(excecao), excecao, excecao.__traceback__))
        )
        with open(ARQUIVO_LOG, 'a', encoding='utf-8') as f:
            f.write(texto)
    except Exception:
        pass  # se nem o log der certo, não tem mais o que fazer


def excecao_nao_tratada(tipo, valor, tb):
    """Captura QUALQUER exceção não tratada em qualquer lugar do app."""
    registrar_erro('excecao_global', valor)
    sys.__excepthook__(tipo, valor, tb)


sys.excepthook = excecao_nao_tratada


KV = '''
<ItemCard@BoxLayout>:
    produto: ''
    quantidade: ''
    endereco: ''
    outros: ''
    orientation: 'vertical'
    size_hint_y: None
    height: self.minimum_height
    padding: dp(10)
    spacing: dp(4)
    canvas.before:
        Color:
            rgba: 0.93, 0.95, 0.98, 1
        RoundedRectangle:
            pos: self.pos
            size: self.size
            radius: [dp(8)]
    BoxLayout:
        size_hint_y: None
        height: self.minimum_height
        Label:
            text: '[b]' + root.produto + '[/b]'
            markup: True
            color: 0.1, 0.1, 0.1, 1
            size_hint_x: 0.65
            halign: 'left'
            valign: 'middle'
            text_size: self.width, None
            size_hint_y: None
            height: self.texture_size[1]
        Label:
            text: 'Qtd: ' + root.quantidade
            color: 0.1, 0.4, 0.1, 1
            size_hint_x: 0.35
            halign: 'right'
            valign: 'middle'
            text_size: self.width, None
            size_hint_y: None
            height: self.texture_size[1]
    Label:
        text: 'Endereço: ' + root.endereco
        color: 0.3, 0.3, 0.3, 1
        halign: 'left'
        valign: 'middle'
        text_size: self.width, None
        size_hint_y: None
        height: self.texture_size[1]
    Label:
        text: root.outros
        color: 0.55, 0.2, 0.2, 1
        halign: 'left'
        valign: 'top'
        text_size: self.width, None
        size_hint_y: None
        height: self.texture_size[1]

MainScreen:

<MainScreen>:
    name: 'main'
    search_input: search_input
    results_box: results_box
    status_label: status_label
    BoxLayout:
        orientation: 'vertical'
        padding: dp(12)
        spacing: dp(10)
        Label:
            text: 'Consulta de Endereço - Estoque'
            font_size: '20sp'
            bold: True
            size_hint_y: None
            height: dp(40)
            color: 0.1, 0.1, 0.1, 1
        BoxLayout:
            size_hint_y: None
            height: dp(48)
            spacing: dp(8)
            TextInput:
                id: search_input
                hint_text: 'Digite o endereço (ex: 1F108C)'
                multiline: False
                font_size: '18sp'
                on_text_validate: root.buscar()
            Button:
                text: 'Buscar'
                size_hint_x: None
                width: dp(90)
                on_release: root.buscar()
            Button:
                text: 'Recarregar'
                size_hint_x: None
                width: dp(100)
                on_release: root.recarregar_manual()
        Label:
            id: status_label
            text: ''
            size_hint_y: None
            height: dp(24)
            color: 0.4, 0.4, 0.4, 1
        ScrollView:
            BoxLayout:
                id: results_box
                orientation: 'vertical'
                size_hint_y: None
                height: self.minimum_height
                spacing: dp(8)
                padding: (0, dp(8))
'''


def normalizar(texto):
    """Remove acentos e deixa minúsculo, pra comparar nomes de coluna/pasta."""
    texto = unicodedata.normalize('NFKD', str(texto))
    texto = texto.encode('ascii', 'ignore').decode('ascii')
    return texto.strip().lower()


PASTAS_PARA_PROCURAR = [
    '/storage/emulated/0/Download',
    '/storage/emulated/0/Documents',
    '/storage/emulated/0/Download/estoque_app/data',
    '/storage/emulated/0',
]

INTERVALO_CHECAGEM_SEGUNDOS = 4


class MainScreen(Screen):
    def on_kv_post(self, base_widget):
        self.atualizar_status()

    def atualizar_status(self, prefixo=''):
        app = App.get_running_app()
        if app.erro_carregamento:
            self.status_label.text = prefixo + app.erro_carregamento
        else:
            nome_arquivo = os.path.basename(app.arquivo_usado) if app.arquivo_usado else '?'
            self.status_label.text = f'{prefixo}{len(app.registros)} registros carregados de {nome_arquivo}'

    def recarregar_manual(self):
        app = App.get_running_app()
        app.carregar_dados()
        self.atualizar_status(prefixo='Atualizado! ')
        if self.search_input.text.strip():
            self.buscar()

    def buscar(self):
        app = App.get_running_app()
        termo = self.search_input.text.strip().upper()
        self.results_box.clear_widgets()

        if not termo:
            self.status_label.text = 'Digite um endereço para buscar.'
            return

        itens = app.endereco_para_itens.get(termo, [])

        if not itens:
            itens = []
            for endereco, lista in app.endereco_para_itens.items():
                if termo in endereco:
                    itens.extend(lista)

        if not itens:
            self.status_label.text = f'Nenhum item encontrado no endereço "{termo}".'
            return

        self.status_label.text = f'{len(itens)} item(ns) encontrado(s) em "{termo}"'

        for produto, armazem, endereco, quantidade in itens:
            outros_enderecos = [
                (e, q) for (p, a, e, q) in app.produto_para_enderecos.get(produto, [])
                if e != endereco
            ]
            if outros_enderecos:
                outros_txt = 'Também está em: ' + ', '.join(
                    f'{e} (qtd {q})' for e, q in outros_enderecos
                )
            else:
                outros_txt = 'Não está em nenhum outro endereço.'

            card = Factory.ItemCard()
            card.produto = produto
            card.quantidade = quantidade
            card.endereco = endereco
            card.outros = outros_txt
            self.results_box.add_widget(card)


def tela_de_erro(mensagem):
    """Tela simples de fallback: mostra o erro na tela em vez de crashar mudo."""
    raiz = BoxLayout(orientation='vertical', padding=dp(16), spacing=dp(10))
    raiz.add_widget(Label(
        text='[b]Ocorreu um erro ao iniciar o app[/b]',
        markup=True,
        color=(0.6, 0.1, 0.1, 1),
        size_hint_y=None,
        height=dp(40),
        font_size='18sp',
    ))
    raiz.add_widget(Label(
        text=f'Detalhes salvos em:\n{ARQUIVO_LOG}',
        color=(0.3, 0.3, 0.3, 1),
        size_hint_y=None,
        height=dp(50),
    ))
    scroll = ScrollView()
    label_erro = Label(
        text=mensagem,
        color=(0.1, 0.1, 0.1, 1),
        size_hint_y=None,
        halign='left',
        valign='top',
    )
    label_erro.bind(texture_size=lambda inst, val: setattr(label_erro, 'height', val[1]))
    label_erro.bind(width=lambda inst, val: setattr(label_erro, 'text_size', (val, None)))
    scroll.add_widget(label_erro)
    raiz.add_widget(scroll)
    return raiz


class EstoqueApp(App):
    def build(self):
        try:
            self.registros = []
            self.endereco_para_itens = defaultdict(list)
            self.produto_para_enderecos = defaultdict(list)
            self.erro_carregamento = ''
            self.arquivo_usado = ''
            self.mtime_usado = None

            self.carregar_dados()

            tela = Builder.load_string(KV)
            Clock.schedule_interval(self.checar_atualizacao_automatica, INTERVALO_CHECAGEM_SEGUNDOS)
            return tela

        except Exception as e:
            # Em vez de deixar o app crashar com tela preta, mostra o erro.
            registrar_erro('build', e)
            traceback_texto = ''.join(traceback.format_exception(type(e), e, e.__traceback__))
            return tela_de_erro(traceback_texto)

    def checar_atualizacao_automatica(self, dt):
        try:
            arquivo_encontrado, _ = self.encontrar_xlsx()

            precisa_recarregar = False
            if arquivo_encontrado and arquivo_encontrado != self.arquivo_usado:
                precisa_recarregar = True
            elif arquivo_encontrado:
                try:
                    mtime_atual = os.path.getmtime(arquivo_encontrado)
                except OSError:
                    mtime_atual = None
                if mtime_atual is not None and mtime_atual != self.mtime_usado:
                    precisa_recarregar = True

            if precisa_recarregar:
                self.carregar_dados()
                tela = self.root
                if tela and hasattr(tela, 'atualizar_status'):
                    tela.atualizar_status(prefixo='Planilha atualizada automaticamente! ')
                    if tela.search_input.text.strip():
                        tela.buscar()
        except Exception as e:
            registrar_erro('checar_atualizacao_automatica', e)

    def encontrar_xlsx(self):
        pastas_vistas = []
        for pasta in PASTAS_PARA_PROCURAR:
            if not os.path.isdir(pasta):
                continue
            pastas_vistas.append(pasta)
            try:
                nomes = os.listdir(pasta)
            except Exception:
                continue
            candidatos = [
                n for n in nomes
                if n.lower().endswith('.xlsx') and not n.startswith('~$')
            ]
            if candidatos:
                candidatos.sort(key=lambda n: 0 if 'estoque' in normalizar(n) else 1)
                return os.path.join(pasta, candidatos[0]), pastas_vistas
        return None, pastas_vistas

    def carregar_dados(self):
        try:
            arquivo_xlsx, pastas_vistas = self.encontrar_xlsx()

            if not arquivo_xlsx:
                self.erro_carregamento = (
                    'Nenhum .xlsx encontrado. Pastas verificadas: ' + ', '.join(pastas_vistas)
                )
                return

            try:
                import openpyxl
            except ImportError:
                self.erro_carregamento = 'Instale a biblioteca "openpyxl" no Pydroid (menu Pip).'
                return

            planilha = openpyxl.load_workbook(arquivo_xlsx, data_only=True)
            aba = None
            for nome_aba in planilha.sheetnames:
                if normalizar(nome_aba) == 'estoque_atual':
                    aba = planilha[nome_aba]
                    break
            if aba is None:
                aba = planilha.active

            linhas = aba.iter_rows(values_only=True)
            cabecalho = next(linhas)
            colunas = {normalizar(c): i for i, c in enumerate(cabecalho) if c is not None}

            idx_produto = colunas.get('produto')
            idx_armazem = colunas.get('armazem')
            idx_endereco = colunas.get('endereco')
            idx_quantidade = colunas.get('quantidade')

            if idx_produto is None or idx_endereco is None:
                self.erro_carregamento = (
                    f'Não encontrei as colunas Produto/Endereco. '
                    f'Colunas achadas: {list(colunas.keys())}'
                )
                return

            novos_registros = []
            novo_endereco_para_itens = defaultdict(list)
            novo_produto_para_enderecos = defaultdict(list)

            for linha in linhas:
                if linha is None:
                    continue

                def pegar(idx):
                    if idx is None or idx >= len(linha):
                        return ''
                    valor = linha[idx]
                    return '' if valor is None else str(valor).strip()

                produto = pegar(idx_produto)
                armazem = pegar(idx_armazem)
                endereco = pegar(idx_endereco).upper()
                quantidade_bruta = linha[idx_quantidade] if idx_quantidade is not None and idx_quantidade < len(linha) else ''

                if not produto or not endereco:
                    continue

                if isinstance(quantidade_bruta, (int, float)):
                    if float(quantidade_bruta).is_integer():
                        quantidade = f'{int(quantidade_bruta)}'
                    else:
                        quantidade = f'{quantidade_bruta:.2f}'.replace('.', ',')
                else:
                    quantidade = str(quantidade_bruta).strip()

                registro = (produto, armazem, endereco, quantidade)
                novos_registros.append(registro)
                novo_endereco_para_itens[endereco].append(registro)
                novo_produto_para_enderecos[produto].append(registro)

            self.registros = novos_registros
            self.endereco_para_itens = novo_endereco_para_itens
            self.produto_para_enderecos = novo_produto_para_enderecos
            self.arquivo_usado = arquivo_xlsx
            self.mtime_usado = os.path.getmtime(arquivo_xlsx)
            self.erro_carregamento = ''

        except Exception as e:
            self.erro_carregamento = f'Erro ao ler a planilha: {e}'
            registrar_erro('carregar_dados', e)


if __name__ == '__main__':
    EstoqueApp().run()
